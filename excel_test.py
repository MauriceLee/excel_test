from openpyxl import load_workbook

wb = load_workbook('input.xlsx') 
ws = wb['Sheet1']

for i in range(2, ws.max_row+1):
    print (ws.cell(i, 1).value, '的身份證字號是', ws.cell(i, 4).value)